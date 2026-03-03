"""Tests for nfe_reader.exporter."""

import csv
import os
import tempfile

import openpyxl
import pytest

from nfe_reader.exporter import NFEExporter
from nfe_reader.parser import NFEParser

FIXTURES_DIR = os.path.join(os.path.dirname(__file__), "fixtures")
SAMPLE_XML = os.path.join(FIXTURES_DIR, "sample_nfe.xml")


@pytest.fixture()
def sample_data():
    return NFEParser().parse_file(SAMPLE_XML)


@pytest.fixture()
def exporter() -> NFEExporter:
    return NFEExporter()


# ---------------------------------------------------------------------------
# export_xlsx (single NF-e)
# ---------------------------------------------------------------------------


class TestExportXlsx:
    def test_creates_file(self, exporter, sample_data, tmp_path):
        out = exporter.export_xlsx(sample_data, str(tmp_path / "output.xlsx"))
        assert os.path.isfile(out)

    def test_appends_extension_if_missing(self, exporter, sample_data, tmp_path):
        out = exporter.export_xlsx(sample_data, str(tmp_path / "output"))
        assert out.endswith(".xlsx")
        assert os.path.isfile(out)

    def test_has_expected_sheets(self, exporter, sample_data, tmp_path):
        out = exporter.export_xlsx(sample_data, str(tmp_path / "output.xlsx"))
        wb = openpyxl.load_workbook(out)
        sheet_names = wb.sheetnames
        for expected in ["Identificação", "Emitente", "Destinatário", "Itens", "Totais", "Protocolo"]:
            assert expected in sheet_names, f"Sheet '{expected}' not found"

    def test_identificacao_sheet_contains_numero(self, exporter, sample_data, tmp_path):
        out = exporter.export_xlsx(sample_data, str(tmp_path / "output.xlsx"))
        wb = openpyxl.load_workbook(out)
        ws = wb["Identificação"]
        values = [str(cell.value) for row in ws.iter_rows() for cell in row if cell.value]
        assert "12341" in values

    def test_emitente_sheet_contains_cnpj(self, exporter, sample_data, tmp_path):
        out = exporter.export_xlsx(sample_data, str(tmp_path / "output.xlsx"))
        wb = openpyxl.load_workbook(out)
        ws = wb["Emitente"]
        values = [str(cell.value) for row in ws.iter_rows() for cell in row if cell.value]
        assert "12345678000195" in values

    def test_itens_sheet_row_count(self, exporter, sample_data, tmp_path):
        out = exporter.export_xlsx(sample_data, str(tmp_path / "output.xlsx"))
        wb = openpyxl.load_workbook(out)
        ws = wb["Itens"]
        # 1 header row + 2 item rows
        assert ws.max_row == 3

    def test_totais_sheet_contains_valor_nota(self, exporter, sample_data, tmp_path):
        out = exporter.export_xlsx(sample_data, str(tmp_path / "output.xlsx"))
        wb = openpyxl.load_workbook(out)
        ws = wb["Totais"]
        values = [str(cell.value) for row in ws.iter_rows() for cell in row if cell.value]
        assert "1100.00" in values

    def test_protocolo_sheet_contains_number(self, exporter, sample_data, tmp_path):
        out = exporter.export_xlsx(sample_data, str(tmp_path / "output.xlsx"))
        wb = openpyxl.load_workbook(out)
        ws = wb["Protocolo"]
        values = [str(cell.value) for row in ws.iter_rows() for cell in row if cell.value]
        assert "135230000123456" in values


# ---------------------------------------------------------------------------
# export_batch_xlsx
# ---------------------------------------------------------------------------


class TestExportBatchXlsx:
    def test_creates_file(self, exporter, sample_data, tmp_path):
        out = exporter.export_batch_xlsx([sample_data], str(tmp_path / "batch.xlsx"))
        assert os.path.isfile(out)

    def test_has_nfe_sheet(self, exporter, sample_data, tmp_path):
        out = exporter.export_batch_xlsx([sample_data], str(tmp_path / "batch.xlsx"))
        wb = openpyxl.load_workbook(out)
        assert "NF-e" in wb.sheetnames

    def test_batch_row_count_matches_records(self, exporter, sample_data, tmp_path):
        records = [sample_data, sample_data]
        out = exporter.export_batch_xlsx(records, str(tmp_path / "batch.xlsx"))
        wb = openpyxl.load_workbook(out)
        ws = wb["NF-e"]
        # 1 header + 2 data rows
        assert ws.max_row == 3

    def test_batch_contains_emitente_nome(self, exporter, sample_data, tmp_path):
        out = exporter.export_batch_xlsx([sample_data], str(tmp_path / "batch.xlsx"))
        wb = openpyxl.load_workbook(out)
        ws = wb["NF-e"]
        values = [str(cell.value) for row in ws.iter_rows() for cell in row if cell.value]
        assert "EMPRESA EMISSORA EXEMPLO LTDA" in values


# ---------------------------------------------------------------------------
# export_csv
# ---------------------------------------------------------------------------


class TestExportCsv:
    def test_creates_file(self, exporter, sample_data, tmp_path):
        out = exporter.export_csv([sample_data], str(tmp_path / "output.csv"))
        assert os.path.isfile(out)

    def test_appends_extension_if_missing(self, exporter, sample_data, tmp_path):
        out = exporter.export_csv([sample_data], str(tmp_path / "output"))
        assert out.endswith(".csv")
        assert os.path.isfile(out)

    def test_csv_has_header(self, exporter, sample_data, tmp_path):
        out = exporter.export_csv([sample_data], str(tmp_path / "output.csv"))
        with open(out, encoding="utf-8-sig") as fh:
            reader = csv.DictReader(fh)
            assert "Chave" in reader.fieldnames
            assert "Valor Nota" in reader.fieldnames

    def test_csv_row_count(self, exporter, sample_data, tmp_path):
        records = [sample_data, sample_data]
        out = exporter.export_csv(records, str(tmp_path / "output.csv"))
        with open(out, encoding="utf-8-sig") as fh:
            rows = list(csv.reader(fh))
        # 1 header + 2 data rows
        assert len(rows) == 3

    def test_csv_contains_cnpj(self, exporter, sample_data, tmp_path):
        out = exporter.export_csv([sample_data], str(tmp_path / "output.csv"))
        with open(out, encoding="utf-8-sig") as fh:
            content = fh.read()
        assert "12345678000195" in content

    def test_csv_contains_valor_nota(self, exporter, sample_data, tmp_path):
        out = exporter.export_csv([sample_data], str(tmp_path / "output.csv"))
        with open(out, encoding="utf-8-sig") as fh:
            content = fh.read()
        assert "1100.00" in content
