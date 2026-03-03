"""
Testes para o módulo nfe_reader.
"""
import os
import tempfile
from decimal import Decimal
from pathlib import Path

import openpyxl
import pytest

from nfe_reader import NFeReader

SAMPLE_XML = Path(__file__).parent / "sample_nfe.xml"


@pytest.fixture
def reader():
    return NFeReader(str(SAMPLE_XML))


# ------------------------------------------------------------------
# Identificação
# ------------------------------------------------------------------

def test_identificacao_numero(reader):
    assert reader.get_identificacao()["numero"] == "1"


def test_identificacao_serie(reader):
    assert reader.get_identificacao()["serie"] == "1"


def test_identificacao_data_emissao(reader):
    assert "2024-01-15" in reader.get_identificacao()["data_emissao"]


def test_identificacao_natureza_operacao(reader):
    assert reader.get_identificacao()["natureza_operacao"] == "VENDA DE PRODUTO"


def test_identificacao_chave(reader):
    chave = reader.get_identificacao()["chave"]
    assert len(chave) == 44


# ------------------------------------------------------------------
# Emitente
# ------------------------------------------------------------------

def test_emitente_cnpj(reader):
    assert reader.get_emitente()["cnpj_cpf"] == "12345678000100"


def test_emitente_nome(reader):
    assert reader.get_emitente()["nome"] == "EMPRESA EXEMPLO LTDA"


def test_emitente_uf(reader):
    assert reader.get_emitente()["uf"] == "SP"


def test_emitente_municipio(reader):
    assert reader.get_emitente()["municipio"] == "SAO PAULO"


# ------------------------------------------------------------------
# Destinatário
# ------------------------------------------------------------------

def test_destinatario_cpf(reader):
    assert reader.get_destinatario()["cnpj_cpf"] == "12345678901"


def test_destinatario_nome(reader):
    assert reader.get_destinatario()["nome"] == "JOAO DA SILVA"


def test_destinatario_uf(reader):
    assert reader.get_destinatario()["uf"] == "SP"


# ------------------------------------------------------------------
# Itens
# ------------------------------------------------------------------

def test_itens_quantidade(reader):
    assert len(reader.get_itens()) == 2


def test_item1_descricao(reader):
    assert reader.get_itens()[0]["descricao"] == "PRODUTO EXEMPLO A"


def test_item1_quantidade(reader):
    assert reader.get_itens()[0]["quantidade"] == Decimal("2.0000")


def test_item1_valor_total(reader):
    assert reader.get_itens()[0]["valor_total"] == Decimal("200.00")


def test_item1_icms(reader):
    assert reader.get_itens()[0]["icms"] == Decimal("24.00")


def test_item1_pis(reader):
    assert reader.get_itens()[0]["pis"] == Decimal("1.30")


def test_item1_cofins(reader):
    assert reader.get_itens()[0]["cofins"] == Decimal("6.00")


def test_item2_desconto(reader):
    assert reader.get_itens()[1]["desconto"] == Decimal("10.00")


def test_item2_ipi(reader):
    assert reader.get_itens()[1]["ipi"] == Decimal("12.00")


# ------------------------------------------------------------------
# Totais
# ------------------------------------------------------------------

def test_totais_valor_nf(reader):
    assert reader.get_totais()["valor_nf"] == Decimal("452.00")


def test_totais_valor_produtos(reader):
    assert reader.get_totais()["valor_produtos"] == Decimal("450.00")


def test_totais_valor_desconto(reader):
    assert reader.get_totais()["valor_desconto"] == Decimal("10.00")


def test_totais_valor_icms(reader):
    assert reader.get_totais()["valor_icms"] == Decimal("52.80")


def test_totais_valor_ipi(reader):
    assert reader.get_totais()["valor_ipi"] == Decimal("12.00")


# ------------------------------------------------------------------
# Exportação xlsx
# ------------------------------------------------------------------

def test_exportar_xlsx_cria_arquivo(reader, tmp_path):
    out = tmp_path / "nfe.xlsx"
    result = reader.exportar_xlsx(str(out))
    assert result == out
    assert out.exists()


def test_exportar_xlsx_abas(reader, tmp_path):
    out = tmp_path / "nfe.xlsx"
    reader.exportar_xlsx(str(out))
    wb = openpyxl.load_workbook(str(out))
    assert "Identificação" in wb.sheetnames
    assert "Emitente" in wb.sheetnames
    assert "Destinatário" in wb.sheetnames
    assert "Itens" in wb.sheetnames
    assert "Totais" in wb.sheetnames


def test_exportar_xlsx_itens_conteudo(reader, tmp_path):
    out = tmp_path / "nfe.xlsx"
    reader.exportar_xlsx(str(out))
    wb = openpyxl.load_workbook(str(out))
    ws = wb["Itens"]
    rows = list(ws.iter_rows(values_only=True))
    # row 0 is header, rows 1 and 2 are items
    assert rows[1][2] == "PRODUTO EXEMPLO A"
    assert rows[2][2] == "PRODUTO EXEMPLO B"


def test_exportar_xlsx_identificacao_conteudo(reader, tmp_path):
    out = tmp_path / "nfe.xlsx"
    reader.exportar_xlsx(str(out))
    wb = openpyxl.load_workbook(str(out))
    ws = wb["Identificação"]
    data = {row[0]: row[1] for row in ws.iter_rows(min_row=2, values_only=True)}
    assert data["Número"] == "1"
    assert data["Série"] == "1"


def test_arquivo_invalido_levanta_excecao():
    with pytest.raises(Exception):
        NFeReader("/tmp/inexistente.xml")


def test_xml_sem_inf_nfe_levanta_excecao(tmp_path):
    bad_xml = tmp_path / "bad.xml"
    bad_xml.write_text("<root><other/></root>")
    with pytest.raises(ValueError, match="infNFe"):
        NFeReader(str(bad_xml))
