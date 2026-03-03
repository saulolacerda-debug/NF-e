"""Export NF-e data to XLSX (and CSV) files using openpyxl."""

from __future__ import annotations

import csv
import os
from typing import List

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .parser import NFEData


# Palette
_COLOR_HEADER = "1F4E79"  # dark blue
_COLOR_SUBHEADER = "2E75B6"  # medium blue
_COLOR_SECTION = "D6E4F0"  # light blue
_COLOR_ALT_ROW = "EBF5FB"  # very light blue
_WHITE = "FFFFFF"


class NFEExporter:
    """Export one or more :class:`~nfe_reader.parser.NFEData` objects.

    Supports:

    * ``export_xlsx`` – export a *single* NF-e to an XLSX workbook with
      separate sheets for each section.
    * ``export_batch_xlsx`` – export a *list* of NF-e records to a single
      XLSX workbook where each row represents one NF-e (summary view).
    * ``export_csv`` – export a *list* of NF-e records to a CSV file
      (summary view).
    """

    # ------------------------------------------------------------------
    # Single-NF-e XLSX
    # ------------------------------------------------------------------

    def export_xlsx(self, data: NFEData, output_path: str) -> str:
        """Write *data* to *output_path* as an XLSX file.

        Returns the absolute path of the created file.
        """
        wb = openpyxl.Workbook()
        wb.remove(wb.active)  # remove default empty sheet

        self._sheet_identificacao(wb, data)
        self._sheet_emitente(wb, data)
        self._sheet_destinatario(wb, data)
        self._sheet_itens(wb, data)
        self._sheet_totais(wb, data)
        self._sheet_protocolo(wb, data)

        output_path = self._ensure_extension(output_path, ".xlsx")
        wb.save(output_path)
        return os.path.abspath(output_path)

    # ------------------------------------------------------------------
    # Batch XLSX (summary)
    # ------------------------------------------------------------------

    def export_batch_xlsx(self, records: List[NFEData], output_path: str) -> str:
        """Write a summary row per NF-e in *records* to *output_path*.

        Returns the absolute path of the created file.
        """
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "NF-e"

        headers = [
            "Chave",
            "Número",
            "Série",
            "Data Emissão",
            "Natureza Operação",
            "Emitente CNPJ/CPF",
            "Emitente Nome",
            "Emitente UF",
            "Destinatário CNPJ/CPF",
            "Destinatário Nome",
            "Destinatário UF",
            "Qtd. Itens",
            "Valor Produtos",
            "Valor Frete",
            "Valor IPI",
            "Valor PIS",
            "Valor COFINS",
            "Valor ICMS",
            "Valor Nota",
            "Protocolo",
            "Status",
        ]
        self._write_header_row(ws, headers, row=1)

        for row_idx, nfe in enumerate(records, start=2):
            fill = PatternFill("solid", fgColor=_COLOR_ALT_ROW) if row_idx % 2 == 0 else None
            values = [
                nfe.identificacao.chave,
                nfe.identificacao.numero,
                nfe.identificacao.serie,
                nfe.identificacao.data_emissao,
                nfe.identificacao.natureza_operacao,
                nfe.emitente.cnpj or nfe.emitente.cpf,
                nfe.emitente.nome,
                nfe.emitente.endereco.uf,
                nfe.destinatario.cnpj or nfe.destinatario.cpf,
                nfe.destinatario.nome,
                nfe.destinatario.endereco.uf,
                len(nfe.itens),
                nfe.totais.valor_produtos,
                nfe.totais.valor_frete,
                nfe.totais.valor_ipi,
                nfe.totais.valor_pis,
                nfe.totais.valor_cofins,
                nfe.totais.valor_icms,
                nfe.totais.valor_nota,
                nfe.protocolo.numero,
                nfe.protocolo.motivo,
            ]
            for col_idx, value in enumerate(values, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.alignment = Alignment(wrap_text=False)
                if fill:
                    cell.fill = fill

        self._auto_width(ws)
        ws.freeze_panes = "A2"

        output_path = self._ensure_extension(output_path, ".xlsx")
        wb.save(output_path)
        return os.path.abspath(output_path)

    # ------------------------------------------------------------------
    # CSV export (summary)
    # ------------------------------------------------------------------

    def export_csv(self, records: List[NFEData], output_path: str) -> str:
        """Write a summary CSV for *records* to *output_path*.

        Returns the absolute path of the created file.
        """
        output_path = self._ensure_extension(output_path, ".csv")
        headers = [
            "Chave",
            "Número",
            "Série",
            "Data Emissão",
            "Natureza Operação",
            "Emitente CNPJ/CPF",
            "Emitente Nome",
            "Emitente UF",
            "Destinatário CNPJ/CPF",
            "Destinatário Nome",
            "Destinatário UF",
            "Qtd. Itens",
            "Valor Produtos",
            "Valor Frete",
            "Valor IPI",
            "Valor PIS",
            "Valor COFINS",
            "Valor ICMS",
            "Valor Nota",
            "Protocolo",
            "Status",
        ]
        with open(output_path, "w", newline="", encoding="utf-8-sig") as fh:
            writer = csv.writer(fh)
            writer.writerow(headers)
            for nfe in records:
                writer.writerow([
                    nfe.identificacao.chave,
                    nfe.identificacao.numero,
                    nfe.identificacao.serie,
                    nfe.identificacao.data_emissao,
                    nfe.identificacao.natureza_operacao,
                    nfe.emitente.cnpj or nfe.emitente.cpf,
                    nfe.emitente.nome,
                    nfe.emitente.endereco.uf,
                    nfe.destinatario.cnpj or nfe.destinatario.cpf,
                    nfe.destinatario.nome,
                    nfe.destinatario.endereco.uf,
                    len(nfe.itens),
                    nfe.totais.valor_produtos,
                    nfe.totais.valor_frete,
                    nfe.totais.valor_ipi,
                    nfe.totais.valor_pis,
                    nfe.totais.valor_cofins,
                    nfe.totais.valor_icms,
                    nfe.totais.valor_nota,
                    nfe.protocolo.numero,
                    nfe.protocolo.motivo,
                ])
        return os.path.abspath(output_path)

    # ------------------------------------------------------------------
    # Sheet builders
    # ------------------------------------------------------------------

    def _sheet_identificacao(self, wb: openpyxl.Workbook, data: NFEData) -> None:
        ws = wb.create_sheet("Identificação")
        self._section_title(ws, "Identificação da NF-e", row=1, ncols=2)
        rows = [
            ("Chave de Acesso", data.identificacao.chave),
            ("Número", data.identificacao.numero),
            ("Série", data.identificacao.serie),
            ("Modelo", data.identificacao.modelo),
            ("Natureza da Operação", data.identificacao.natureza_operacao),
            ("Data de Emissão", data.identificacao.data_emissao),
            ("Data Saída/Entrada", data.identificacao.data_saida_entrada),
            ("Tipo de Operação", data.identificacao.tipo_operacao),
            ("UF", data.identificacao.uf),
            ("Município Fato Gerador", data.identificacao.municipio_fato_gerador),
            ("Finalidade", data.identificacao.finalidade),
            ("Destino", data.identificacao.destino),
            ("Informações Complementares", data.informacoes_complementares),
        ]
        self._write_key_value_rows(ws, rows, start_row=2)
        self._auto_width(ws)

    def _sheet_emitente(self, wb: openpyxl.Workbook, data: NFEData) -> None:
        ws = wb.create_sheet("Emitente")
        self._section_title(ws, "Dados do Emitente", row=1, ncols=2)
        em = data.emitente
        rows = [
            ("CNPJ", em.cnpj),
            ("CPF", em.cpf),
            ("Nome / Razão Social", em.nome),
            ("Nome Fantasia", em.fantasia),
            ("Inscrição Estadual", em.ie),
            ("CRT", em.crt),
            ("Logradouro", em.endereco.logradouro),
            ("Número", em.endereco.numero),
            ("Complemento", em.endereco.complemento),
            ("Bairro", em.endereco.bairro),
            ("Município", em.endereco.municipio),
            ("UF", em.endereco.uf),
            ("CEP", em.endereco.cep),
            ("País", em.endereco.pais),
            ("Telefone", em.endereco.telefone),
        ]
        self._write_key_value_rows(ws, rows, start_row=2)
        self._auto_width(ws)

    def _sheet_destinatario(self, wb: openpyxl.Workbook, data: NFEData) -> None:
        ws = wb.create_sheet("Destinatário")
        self._section_title(ws, "Dados do Destinatário", row=1, ncols=2)
        d = data.destinatario
        rows = [
            ("CNPJ", d.cnpj),
            ("CPF", d.cpf),
            ("Nome / Razão Social", d.nome),
            ("Inscrição Estadual", d.ie),
            ("E-mail", d.email),
            ("Logradouro", d.endereco.logradouro),
            ("Número", d.endereco.numero),
            ("Complemento", d.endereco.complemento),
            ("Bairro", d.endereco.bairro),
            ("Município", d.endereco.municipio),
            ("UF", d.endereco.uf),
            ("CEP", d.endereco.cep),
            ("País", d.endereco.pais),
            ("Telefone", d.endereco.telefone),
        ]
        self._write_key_value_rows(ws, rows, start_row=2)
        self._auto_width(ws)

    def _sheet_itens(self, wb: openpyxl.Workbook, data: NFEData) -> None:
        ws = wb.create_sheet("Itens")
        headers = [
            "Nº Item",
            "Código",
            "Descrição",
            "NCM",
            "CFOP",
            "Unid.",
            "Quantidade",
            "Valor Unit. (R$)",
            "Valor Total (R$)",
            "ICMS (R$)",
            "IPI (R$)",
            "PIS (R$)",
            "COFINS (R$)",
        ]
        self._write_header_row(ws, headers, row=1)
        for row_idx, item in enumerate(data.itens, start=2):
            fill = PatternFill("solid", fgColor=_COLOR_ALT_ROW) if row_idx % 2 == 0 else None
            values = [
                item.numero,
                item.codigo,
                item.descricao,
                item.ncm,
                item.cfop,
                item.unidade,
                item.quantidade,
                item.valor_unitario,
                item.valor_total,
                item.icms,
                item.ipi,
                item.pis,
                item.cofins,
            ]
            for col_idx, value in enumerate(values, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                if fill:
                    cell.fill = fill
        self._auto_width(ws)
        ws.freeze_panes = "A2"

    def _sheet_totais(self, wb: openpyxl.Workbook, data: NFEData) -> None:
        ws = wb.create_sheet("Totais")
        self._section_title(ws, "Totais da NF-e", row=1, ncols=2)
        t = data.totais
        rows = [
            ("Base de Cálculo ICMS (R$)", t.base_calculo_icms),
            ("Valor ICMS (R$)", t.valor_icms),
            ("Valor ICMS Desonerado (R$)", t.valor_icms_desonerado),
            ("Valor FCP (R$)", t.valor_fcp),
            ("Base de Cálculo ST (R$)", t.valor_bcst),
            ("Valor ST (R$)", t.valor_st),
            ("Valor FCP ST (R$)", t.valor_fcp_st),
            ("Valor Produtos (R$)", t.valor_produtos),
            ("Valor Frete (R$)", t.valor_frete),
            ("Valor Seguro (R$)", t.valor_seguro),
            ("Outros (R$)", t.outros),
            ("Valor IPI (R$)", t.valor_ipi),
            ("Valor PIS (R$)", t.valor_pis),
            ("Valor COFINS (R$)", t.valor_cofins),
            ("Valor Total da Nota (R$)", t.valor_nota),
        ]
        self._write_key_value_rows(ws, rows, start_row=2)
        self._auto_width(ws)

    def _sheet_protocolo(self, wb: openpyxl.Workbook, data: NFEData) -> None:
        ws = wb.create_sheet("Protocolo")
        self._section_title(ws, "Protocolo de Autorização", row=1, ncols=2)
        p = data.protocolo
        rows = [
            ("Número do Protocolo", p.numero),
            ("Data de Autorização", p.data_autorizacao),
            ("Código Status", p.codigo_status),
            ("Motivo", p.motivo),
            ("Digest Value", p.digest_value),
        ]
        self._write_key_value_rows(ws, rows, start_row=2)
        self._auto_width(ws)

    # ------------------------------------------------------------------
    # Styling helpers
    # ------------------------------------------------------------------

    def _write_header_row(
        self, ws: openpyxl.worksheet.worksheet.Worksheet, headers: List[str], row: int
    ) -> None:
        header_font = Font(bold=True, color=_WHITE)
        header_fill = PatternFill("solid", fgColor=_COLOR_HEADER)
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.row_dimensions[row].height = 30

    def _section_title(
        self,
        ws: openpyxl.worksheet.worksheet.Worksheet,
        title: str,
        row: int,
        ncols: int,
    ) -> None:
        ws.merge_cells(
            start_row=row,
            start_column=1,
            end_row=row,
            end_column=ncols,
        )
        cell = ws.cell(row=row, column=1, value=title)
        cell.font = Font(bold=True, size=12, color=_WHITE)
        cell.fill = PatternFill("solid", fgColor=_COLOR_SUBHEADER)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[row].height = 25

    def _write_key_value_rows(
        self,
        ws: openpyxl.worksheet.worksheet.Worksheet,
        rows: list,
        start_row: int,
    ) -> None:
        label_font = Font(bold=True)
        section_fill = PatternFill("solid", fgColor=_COLOR_SECTION)
        alt_fill = PatternFill("solid", fgColor=_COLOR_ALT_ROW)
        for offset, (key, value) in enumerate(rows):
            r = start_row + offset
            label_cell = ws.cell(row=r, column=1, value=key)
            label_cell.font = label_font
            label_cell.fill = section_fill
            value_cell = ws.cell(row=r, column=2, value=value)
            value_cell.fill = alt_fill if offset % 2 == 0 else PatternFill()
            value_cell.alignment = Alignment(wrap_text=True)

    def _auto_width(self, ws: openpyxl.worksheet.worksheet.Worksheet) -> None:
        for col_cells in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col_cells[0].column)
            for cell in col_cells:
                try:
                    cell_len = len(str(cell.value)) if cell.value is not None else 0
                    if cell_len > max_len:
                        max_len = cell_len
                except (TypeError, AttributeError):
                    pass
            ws.column_dimensions[col_letter].width = min(max_len + 4, 60)

    # ------------------------------------------------------------------
    # Utility
    # ------------------------------------------------------------------

    @staticmethod
    def _ensure_extension(path: str, ext: str) -> str:
        """Append *ext* to *path* if it doesn't already end with it."""
        if not path.lower().endswith(ext):
            return path + ext
        return path
