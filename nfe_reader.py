"""
Leitor de NF-e (Nota Fiscal Eletrônica) com exportação para xlsx.
"""
import logging
import xml.etree.ElementTree as ET
from decimal import Decimal, InvalidOperation
from pathlib import Path

logger = logging.getLogger(__name__)

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

_NS = "http://www.portalfiscal.inf.br/nfe"


def _tag(name: str) -> str:
    return f"{{{_NS}}}{name}"


def _find_text(element, path: str, default: str = "") -> str:
    parts = path.split("/")
    current = element
    for part in parts:
        if current is None:
            return default
        current = current.find(_tag(part))
    if current is None or current.text is None:
        return default
    return current.text.strip()


def _decimal(value: str) -> Decimal:
    try:
        return Decimal(value)
    except (InvalidOperation, TypeError):
        logger.warning("Valor decimal inválido: %r — substituído por zero.", value)
        return Decimal("0")


class NFeReader:
    """Lê um arquivo XML de NF-e e disponibiliza seus dados."""

    def __init__(self, xml_path: str):
        self._path = Path(xml_path)
        tree = ET.parse(self._path)
        root = tree.getroot()
        # Handle both bare NFe and nfeProc wrappers
        self._inf_nfe = (
            root.find(f".//{_tag('infNFe')}")
        )
        if self._inf_nfe is None:
            raise ValueError(f"Elemento infNFe não encontrado em {xml_path}")

    # ------------------------------------------------------------------
    # Public data accessors
    # ------------------------------------------------------------------

    def get_identificacao(self) -> dict:
        """Dados de identificação da NF-e."""
        ide = self._inf_nfe.find(_tag("ide"))
        return {
            "numero": _find_text(ide, "nNF"),
            "serie": _find_text(ide, "serie"),
            "data_emissao": _find_text(ide, "dhEmi") or _find_text(ide, "dEmi"),
            "natureza_operacao": _find_text(ide, "natOp"),
            "tipo_operacao": _find_text(ide, "tpOper") or _find_text(ide, "tpNF"),
            "chave": self._inf_nfe.get("Id", "").replace("NFe", ""),
        }

    def get_emitente(self) -> dict:
        """Dados do emitente."""
        emit = self._inf_nfe.find(_tag("emit"))
        cnpj = _find_text(emit, "CNPJ")
        cpf = _find_text(emit, "CPF")
        return {
            "cnpj_cpf": cnpj or cpf,
            "nome": _find_text(emit, "xNome"),
            "nome_fantasia": _find_text(emit, "xFant"),
            "ie": _find_text(emit, "IE"),
            "logradouro": _find_text(emit, "enderEmit/xLgr"),
            "numero": _find_text(emit, "enderEmit/nro"),
            "bairro": _find_text(emit, "enderEmit/xBairro"),
            "municipio": _find_text(emit, "enderEmit/xMun"),
            "uf": _find_text(emit, "enderEmit/UF"),
            "cep": _find_text(emit, "enderEmit/CEP"),
        }

    def get_destinatario(self) -> dict:
        """Dados do destinatário."""
        dest = self._inf_nfe.find(_tag("dest"))
        if dest is None:
            return {}
        cnpj = _find_text(dest, "CNPJ")
        cpf = _find_text(dest, "CPF")
        return {
            "cnpj_cpf": cnpj or cpf,
            "nome": _find_text(dest, "xNome"),
            "ie": _find_text(dest, "IE"),
            "logradouro": _find_text(dest, "enderDest/xLgr"),
            "numero": _find_text(dest, "enderDest/nro"),
            "bairro": _find_text(dest, "enderDest/xBairro"),
            "municipio": _find_text(dest, "enderDest/xMun"),
            "uf": _find_text(dest, "enderDest/UF"),
            "cep": _find_text(dest, "enderDest/CEP"),
        }

    def get_itens(self) -> list[dict]:
        """Lista de itens/produtos da NF-e."""
        itens = []
        for det in self._inf_nfe.findall(_tag("det")):
            prod = det.find(_tag("prod"))
            imposto = det.find(_tag("imposto"))
            item = {
                "numero": det.get("nItem", ""),
                "codigo": _find_text(prod, "cProd"),
                "descricao": _find_text(prod, "xProd"),
                "ncm": _find_text(prod, "NCM"),
                "cfop": _find_text(prod, "CFOP"),
                "unidade": _find_text(prod, "uCom"),
                "quantidade": _decimal(_find_text(prod, "qCom")),
                "valor_unitario": _decimal(_find_text(prod, "vUnCom")),
                "valor_total": _decimal(_find_text(prod, "vProd")),
                "desconto": _decimal(_find_text(prod, "vDesc")),
                "icms": _get_imposto_valor(imposto, "ICMS", "vICMS"),
                "ipi": _get_imposto_valor(imposto, "IPI", "vIPI"),
                "pis": _get_imposto_valor(imposto, "PIS", "vPIS"),
                "cofins": _get_imposto_valor(imposto, "COFINS", "vCOFINS"),
            }
            itens.append(item)
        return itens

    def get_totais(self) -> dict:
        """Totais da NF-e."""
        icms_tot = self._inf_nfe.find(f".//{_tag('ICMSTot')}")
        return {
            "valor_produtos": _decimal(_find_text(icms_tot, "vProd")),
            "valor_frete": _decimal(_find_text(icms_tot, "vFrete")),
            "valor_seguro": _decimal(_find_text(icms_tot, "vSeg")),
            "valor_desconto": _decimal(_find_text(icms_tot, "vDesc")),
            "valor_outros": _decimal(_find_text(icms_tot, "vOutro")),
            "valor_icms": _decimal(_find_text(icms_tot, "vICMS")),
            "valor_ipi": _decimal(_find_text(icms_tot, "vIPI")),
            "valor_pis": _decimal(_find_text(icms_tot, "vPIS")),
            "valor_cofins": _decimal(_find_text(icms_tot, "vCOFINS")),
            "valor_nf": _decimal(_find_text(icms_tot, "vNF")),
        }

    # ------------------------------------------------------------------
    # xlsx export
    # ------------------------------------------------------------------

    def exportar_xlsx(self, output_path: str) -> Path:
        """Exporta os dados da NF-e para um arquivo xlsx.

        Args:
            output_path: caminho do arquivo xlsx a ser gerado.

        Returns:
            Path do arquivo gerado.
        """
        wb = openpyxl.Workbook()

        _build_sheet_identificacao(wb, self.get_identificacao())
        _build_sheet_emitente(wb, self.get_emitente())
        _build_sheet_destinatario(wb, self.get_destinatario())
        _build_sheet_itens(wb, self.get_itens())
        _build_sheet_totais(wb, self.get_totais())

        # Remove the default empty sheet
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

        out = Path(output_path)
        wb.save(out)
        return out


# ------------------------------------------------------------------
# Internal helpers
# ------------------------------------------------------------------

def _get_imposto_valor(imposto, grupo: str, campo: str) -> Decimal:
    """Busca o valor de um imposto dentro de qualquer sub-tag de um grupo.

    Args:
        imposto: elemento XML ``imposto`` do item (``xml.etree.ElementTree.Element``
            ou ``None``).
        grupo: nome da tag de grupo do imposto, ex.: ``"ICMS"``, ``"PIS"``.
        campo: nome da tag com o valor monetário, ex.: ``"vICMS"``, ``"vPIS"``.

    Returns:
        Valor decimal do imposto, ou ``Decimal("0")`` quando não encontrado.
    """
    if imposto is None:
        return Decimal("0")
    grupo_el = imposto.find(_tag(grupo))
    if grupo_el is None:
        return Decimal("0")
    for child in grupo_el:
        val_el = child.find(_tag(campo))
        if val_el is not None and val_el.text:
            return _decimal(val_el.text)
    return Decimal("0")


_HEADER_FILL = PatternFill(fill_type="solid", fgColor="1F4E79")
_HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
_LABEL_FONT = Font(bold=True)


def _header(ws, col_a_label: str = "Campo", col_b_label: str = "Valor"):
    ws.append([col_a_label, col_b_label])
    for cell in ws[1]:
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center")


def _auto_width(ws):
    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=0)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 80)


def _build_sheet_identificacao(wb, dados: dict):
    ws = wb.create_sheet("Identificação")
    _header(ws)
    labels = {
        "numero": "Número",
        "serie": "Série",
        "data_emissao": "Data de Emissão",
        "natureza_operacao": "Natureza da Operação",
        "tipo_operacao": "Tipo de Operação",
        "chave": "Chave de Acesso",
    }
    for key, label in labels.items():
        ws.append([label, dados.get(key, "")])
        ws.cell(row=ws.max_row, column=1).font = _LABEL_FONT
    _auto_width(ws)


def _build_sheet_emitente(wb, dados: dict):
    ws = wb.create_sheet("Emitente")
    _header(ws)
    labels = {
        "cnpj_cpf": "CNPJ/CPF",
        "nome": "Nome / Razão Social",
        "nome_fantasia": "Nome Fantasia",
        "ie": "Inscrição Estadual",
        "logradouro": "Logradouro",
        "numero": "Número",
        "bairro": "Bairro",
        "municipio": "Município",
        "uf": "UF",
        "cep": "CEP",
    }
    for key, label in labels.items():
        ws.append([label, dados.get(key, "")])
        ws.cell(row=ws.max_row, column=1).font = _LABEL_FONT
    _auto_width(ws)


def _build_sheet_destinatario(wb, dados: dict):
    ws = wb.create_sheet("Destinatário")
    _header(ws)
    labels = {
        "cnpj_cpf": "CNPJ/CPF",
        "nome": "Nome / Razão Social",
        "ie": "Inscrição Estadual",
        "logradouro": "Logradouro",
        "numero": "Número",
        "bairro": "Bairro",
        "municipio": "Município",
        "uf": "UF",
        "cep": "CEP",
    }
    for key, label in labels.items():
        ws.append([label, dados.get(key, "")])
        ws.cell(row=ws.max_row, column=1).font = _LABEL_FONT
    _auto_width(ws)


def _build_sheet_itens(wb, itens: list[dict]):
    ws = wb.create_sheet("Itens")
    headers = [
        "Nº", "Código", "Descrição", "NCM", "CFOP", "Unidade",
        "Quantidade", "Valor Unitário", "Valor Total", "Desconto",
        "ICMS", "IPI", "PIS", "COFINS",
    ]
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center")
    for item in itens:
        ws.append([
            item["numero"],
            item["codigo"],
            item["descricao"],
            item["ncm"],
            item["cfop"],
            item["unidade"],
            float(item["quantidade"]),
            float(item["valor_unitario"]),
            float(item["valor_total"]),
            float(item["desconto"]),
            float(item["icms"]),
            float(item["ipi"]),
            float(item["pis"]),
            float(item["cofins"]),
        ])
    _auto_width(ws)


def _build_sheet_totais(wb, totais: dict):
    ws = wb.create_sheet("Totais")
    _header(ws)
    labels = {
        "valor_produtos": "Valor dos Produtos",
        "valor_frete": "Valor do Frete",
        "valor_seguro": "Valor do Seguro",
        "valor_desconto": "Valor do Desconto",
        "valor_outros": "Outros",
        "valor_icms": "ICMS",
        "valor_ipi": "IPI",
        "valor_pis": "PIS",
        "valor_cofins": "COFINS",
        "valor_nf": "Valor Total da NF-e",
    }
    for key, label in labels.items():
        ws.append([label, float(totais.get(key, 0))])
        ws.cell(row=ws.max_row, column=1).font = _LABEL_FONT
    _auto_width(ws)
